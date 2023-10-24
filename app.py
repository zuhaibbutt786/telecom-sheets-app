import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO

def highlight_open_1(df, sheet_name):
    try:
        sheet = df[sheet_name]
        for col_name in sheet.columns:
            column = sheet[col_name]
            open1_indices = column.index[column == 'Open[1]'].tolist()

            if open1_indices:
                for index in open1_indices:
                    sheet = writer.sheets[sheet_name]
                    cell = sheet.cell(row=index + 2, column=col_name + 1)  # Add 2 because Excel is 1-indexed
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    except KeyError:
        # Handle the case when the sheet is not found or not visible
        pass

    return df

def main():
    st.title("Excel Highlighter")

    # Upload a file
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

    if uploaded_file is not None:
        # Read the Excel file
        df = pd.read_excel(uploaded_file, sheet_name=None)

        # Show the list of sheets in the uploaded file
        sheet_names = list(df.keys())
        selected_sheet = st.selectbox("Select a sheet to process:", sheet_names)

        if st.button("Highlight 'Open[1]'"):
            # Highlight 'Open[1]' in the selected sheet
            df = highlight_open_1(df, selected_sheet)

            # Save the highlighted data to a new Excel file
            new_filename = f"highlighted_{selected_sheet}.xlsx"
            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    for sheet_name in df:
                        df[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
                        if sheet_name == selected_sheet:
                            highlight_open_1(writer, sheet_name)
                    writer.save()
                buffer.seek(0)
                st.download_button("Download Highlighted Excel File", buffer)

    if uploaded_file is None:
        st.warning("Please upload an Excel file.")

if __name__ == "__main__":
    main()
